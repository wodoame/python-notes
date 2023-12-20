# Here's a basic tutorial on how to use `openpyxl` to work with Excel files in Python.
# We'll cover reading and writing data, as well as some common operations.
# Make sure you have `openpyxl` installed:
# pip install openpyxl

# Now, let's get started with a simple example:

### Reading from an Excel file:
import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select the active sheet
sheet = workbook.active

# Accessing cell values
cell_value = sheet['A1'].value
print(f'Cell A1: {cell_value}')

# Iterating through rows
for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3, values_only=True):
    print(row)

# In this example, we load an Excel file named 'example.xlsx', access the active sheet, and retrieve values from specific cells.
# The `iter_rows` method is used to iterate through a range of rows and columns.

### Writing to an Excel file:

import openpyxl

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Writing data to cells
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'

# Save the workbook
workbook.save('output.xlsx')


# Here, we create a new workbook, access the active sheet, and write data to specific cells.
# Finally, we save the workbook with the name 'output.xlsx'.

### Working with Styles:

import openpyxl
from openpyxl.styles import Font, Alignment

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Styling cells
sheet['A1'].font = Font(bold=True, color='FF0000')  # Bold red text
sheet['A1'].alignment = Alignment(horizontal='center')  # Centered text

# Save the workbook
workbook.save('styled_output.xlsx')


# This example demonstrates how to apply styles to cells using the `Font` and `Alignment` classes from `openpyxl.styles`.
### Handling Formulas:


import openpyxl

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Writing formulas to cells
sheet['A1'] = 10
sheet['B1'] = 20
sheet['C1'] = '=A1 + B1'  # Formula in C1

# Save the workbook
workbook.save('formulas_output.xlsx')

# I was trying to convert excel files to csv and I got this code example
# To convert an Excel file to CSV using `openpyxl`, you can follow these steps:

# 1. Use the following Python script to convert an Excel file to CSV:

   import openpyxl
   import csv

   def excel_to_csv(input_excel_file, output_csv_file):
       # Load the Excel workbook
       workbook = openpyxl.load_workbook(input_excel_file)

       # Choose the active sheet or specify the sheet name if needed
       sheet = workbook.active

       # Open a CSV file for writing
       with open(output_csv_file, 'w', newline='') as csv_file:
           csv_writer = csv.writer(csv_file)

           # Iterate through rows in the sheet and write to CSV
           for row in sheet.iter_rows(values_only=True):
               csv_writer.writerow(row)

   # Example usage
   input_excel_file = 'input.xlsx'
   output_csv_file = 'output.csv'
   excel_to_csv(input_excel_file, output_csv_file)
   

# Replace `'input.xlsx'` with the path to your Excel file and `'output.csv'` with the desired path for the CSV output file.
# This script uses the `openpyxl` library to load the Excel workbook, accesses the active sheet, and iterates through the rows to write them to a CSV file using the `csv` module.
# Note that this script assumes that the Excel file has a single active sheet, and it ignores any formatting or formulas in the Excel file.

# Make sure to adapt the script to your specific needs, especially if you have multiple sheets or need to handle specific Excel features.
# You can write formulas to cells by assigning a string containing the formula to the cell.

# These examples cover some of the basic operations with `openpyxl`.
# For more detailed information, refer to the official `openpyxl` documentation: [openpyxl Documentation](https://openpyxl.readthedocs.io/).
